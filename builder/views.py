import json
import uuid
from io import BytesIO

from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.conf import settings
from django.core.mail import send_mail
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.views.decorators.http import require_http_methods

from .models import BuilderRun, PasswordResetToken
from .ai_engine import chat_with_ai, generate_greeting, get_suggestions_for_message

FREE_LIMIT = getattr(settings, 'FREE_RUN_LIMIT', 3)


# ── Auth ──────────────────────────────────────────────────────────────────────

def login_view(request):
    if request.user.is_authenticated:
        return redirect('index')
    error = None
    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '')
        user = authenticate(request, username=username, password=password)
        if user:
            login(request, user)
            return redirect(request.GET.get('next', '/'))
        error = 'Invalid username or password.'
    return render(request, 'builder/login.html', {'error': error})


def signup_view(request):
    if request.user.is_authenticated:
        return redirect('index')
    error = None
    if request.method == 'POST':
        username  = request.POST.get('username', '').strip()
        email     = request.POST.get('email', '').strip()
        password1 = request.POST.get('password1', '')
        password2 = request.POST.get('password2', '')
        if not username or not password1:
            error = 'Username and password are required.'
        elif password1 != password2:
            error = 'Passwords do not match.'
        elif len(password1) < 8:
            error = 'Password must be at least 8 characters.'
        elif User.objects.filter(username=username).exists():
            error = 'That username is already taken.'
        else:
            user = User.objects.create_user(username=username, email=email, password=password1)
            login(request, user)
            return redirect('/')
    return render(request, 'builder/signup.html', {'error': error})


def logout_view(request):
    logout(request)
    return redirect('/login/')


# ── Custom password reset (Bundle — not Django auth URLs) ─────────────────────

def _user_by_email(email):
    """Match account by email or username (many Bundle users sign in with email as username)."""
    email = (email or '').strip()
    if not email:
        return None
    user = User.objects.filter(email__iexact=email).first()
    if user:
        return user
    return User.objects.filter(username__iexact=email).first()


def _password_reset_link(request, token_uuid):
    """Absolute reset URL — uses current host/port locally and your domain when deployed."""
    path = reverse('bundle_password_reset_confirm', kwargs={'token': token_uuid})
    if getattr(settings, 'SITE_URL', ''):
        return f"{settings.SITE_URL}{path}"
    return request.build_absolute_uri(path)


def _send_password_reset_email(request, user, reset_url):
    subject = 'Bundle — Reset your password'
    message = (
        f"Hi{' ' + user.first_name if user.first_name else ''},\n\n"
        "We received a request to reset your Bundle password.\n\n"
        f"Set a new password using this link (expires in "
        f"{getattr(settings, 'PASSWORD_RESET_TIMEOUT_HOURS', 24)} hours):\n\n"
        f"{reset_url}\n\n"
        "If you did not request this, you can ignore this email.\n\n"
        "— Bundle"
    )
    send_mail(
        subject,
        message,
        settings.DEFAULT_FROM_EMAIL,
        [user.email or user.username],
        fail_silently=False,
    )


@require_http_methods(['GET', 'POST'])
def password_reset_request(request):
    if request.user.is_authenticated:
        return redirect('index')

    success = False
    error = None

    if request.method == 'POST':
        email = request.POST.get('email', '').strip()
        if not email:
            error = 'Please enter your email address.'
        else:
            user = _user_by_email(email)
            if user:
                PasswordResetToken.objects.filter(user=user, used=False).update(used=True)
                reset_token = PasswordResetToken.objects.create(user=user)
                reset_url = _password_reset_link(request, reset_token.token)
                try:
                    _send_password_reset_email(request, user, reset_url)
                except Exception:
                    error = (
                        'We could not send the email right now. '
                        'Please try again in a few minutes.'
                    )
                else:
                    success = True
            else:
                # Do not reveal whether the email is registered.
                success = True

    return render(request, 'builder/password_reset_request.html', {
        'error': error,
        'success': success,
    })


@require_http_methods(['GET', 'POST'])
def password_reset_confirm(request, token):
    if request.user.is_authenticated:
        return redirect('index')

    try:
        token_uuid = uuid.UUID(str(token))
    except (ValueError, AttributeError):
        return render(request, 'builder/password_reset_confirm.html', {
            'invalid': True,
            'error': None,
        })

    reset_token = PasswordResetToken.objects.filter(token=token_uuid).select_related('user').first()
    if not reset_token or not reset_token.is_valid:
        return render(request, 'builder/password_reset_confirm.html', {
            'invalid': True,
            'error': None,
        })

    error = None
    if request.method == 'POST':
        password1 = request.POST.get('password1', '')
        password2 = request.POST.get('password2', '')
        if len(password1) < 8:
            error = 'Password must be at least 8 characters.'
        elif password1 != password2:
            error = 'Passwords do not match.'
        else:
            user = reset_token.user
            user.set_password(password1)
            user.save(update_fields=['password'])
            reset_token.mark_used()
            PasswordResetToken.objects.filter(user=user, used=False).update(used=True)
            return redirect('login')

    return render(request, 'builder/password_reset_confirm.html', {
        'invalid': False,
        'error': error,
        'token': token_uuid,
    })


# ── Main app ──────────────────────────────────────────────────────────────────

def _dashboard_context(request):
    runs_used = BuilderRun.objects.filter(user=request.user).count()
    return {
        'threads':    BuilderRun.objects.filter(user=request.user).order_by('-updated_at')[:20],
        'runs_used':  runs_used,
        'free_limit': FREE_LIMIT,
        'is_locked':  runs_used >= FREE_LIMIT,
    }


@login_required
def index(request):
    return render(request, 'builder/dashboard.html', _dashboard_context(request))


@login_required
def chat_new(request):
    ctx = _dashboard_context(request)
    if ctx['is_locked']:
        return redirect('index')
    return render(request, 'builder/chat.html', {
        **ctx,
        'chat_mode': 'new',
        'run_id':    None,
    })


@login_required
def chat_thread(request, run_id):
    run = get_object_or_404(BuilderRun, id=run_id, user=request.user)
    request.session['run_id'] = run.id
    return render(request, 'builder/chat.html', {
        **_dashboard_context(request),
        'chat_mode': 'thread',
        'run_id':    run.id,
    })


def start_chat(request):
    if not request.user.is_authenticated:
        return JsonResponse({'type': 'unauthenticated'})

    run_id = request.session.get('run_id')
    if run_id:
        try:
            run     = BuilderRun.objects.get(id=run_id, user=request.user)
            history = run.get_chat_history()
            if history:
                last_ai = next(
                    (m['content'] for m in reversed(history) if m.get('role') in ('assistant', 'ai')),
                    '',
                )
                suggestions = []
                if run.status not in ('plan_generated', 'cta_clicked'):
                    suggestions = get_suggestions_for_message(last_ai)
                return JsonResponse({
                    'type':        'resume',
                    'history':     history,
                    'stage':       run.current_stage,
                    'status':      run.status,
                    'run_id':      run.id,
                    'suggestions': suggestions,
                })
        except BuilderRun.DoesNotExist:
            pass

    runs_used = BuilderRun.objects.filter(user=request.user).count()
    return JsonResponse({
        'type':       'new',
        'runs_used':  runs_used,
        'free_limit': FREE_LIMIT,
        'locked':     runs_used >= FREE_LIMIT,
    })


@login_required
@require_http_methods(["POST"])
def new_run(request):
    runs_used = BuilderRun.objects.filter(user=request.user).count()
    if runs_used >= FREE_LIMIT:
        return JsonResponse(
            {'error': 'free_limit_reached', 'runs_used': runs_used, 'free_limit': FREE_LIMIT},
            status=403,
        )

    if not request.session.session_key:
        request.session.create()

    run = BuilderRun.objects.create(
        user=request.user,
        session_key=request.session.session_key or '',
        current_stage=1,
    )
    request.session['run_id'] = run.id

    try:
        greeting, suggestions = generate_greeting(run)
    except Exception:
        greeting = (
            "Welcome! I'm the Bundle AI Training Builder. I'll help you build a specific, "
            "role-by-role training plan for your team.\n\n"
            "To get started — what's your company name, industry, and approximately how many people are on your team?"
        )
        from .ai_engine import STAGE_SUGGESTIONS
        suggestions = STAGE_SUGGESTIONS['q_company']
        history = [{"role": "assistant", "content": greeting}]
        run.set_chat_history(history)
        run.save()

    return JsonResponse({
        'run_id':     run.id,
        'message':    greeting,
        'suggestions': suggestions,
        'stage':      1,
        'status':     'in_progress',
        'runs_used':  runs_used + 1,
        'free_limit': FREE_LIMIT,
    })


@login_required
@require_http_methods(["POST"])
def chat(request):
    try:
        data    = json.loads(request.body)
        message = data.get('message', '').strip()
        run_id  = data.get('run_id')

        if not message:
            return JsonResponse({'error': 'Message is required.'}, status=400)
        if not run_id:
            return JsonResponse({'error': 'Run ID is required.'}, status=400)

        run = get_object_or_404(BuilderRun, id=run_id, user=request.user)
        ai_reply, suggestions = chat_with_ai(run, message)

        return JsonResponse({
            'message':     ai_reply,
            'suggestions': suggestions,
            'stage':       run.current_stage,
            'status':      run.status,
            'run_id':      run.id,
        })
    except Exception as e:
        return JsonResponse({'error': 'Something went wrong. Please try again.'}, status=500)


@login_required
@require_http_methods(["POST"])
def upload_file(request, run_id):
    run = get_object_or_404(BuilderRun, id=run_id, user=request.user)
    f   = request.FILES.get('file')
    if not f:
        return JsonResponse({'error': 'No file provided.'}, status=400)

    filename = f.name
    ext      = filename.rsplit('.', 1)[-1].lower() if '.' in filename else ''
    raw      = f.read()

    try:
        if ext in ('txt', 'csv'):
            text = raw.decode('utf-8', errors='replace')
        elif ext == 'pdf':
            from pypdf import PdfReader
            reader = PdfReader(BytesIO(raw))
            text   = '\n'.join(page.extract_text() or '' for page in reader.pages)
        elif ext == 'docx':
            from docx import Document
            doc  = Document(BytesIO(raw))
            text = '\n'.join(p.text for p in doc.paragraphs)
        elif ext in ('xlsx', 'xls'):
            from openpyxl import load_workbook
            wb    = load_workbook(BytesIO(raw), read_only=True, data_only=True)
            lines = []
            for ws in wb.worksheets:
                for row in ws.iter_rows(values_only=True):
                    lines.append(', '.join(str(c) if c is not None else '' for c in row))
            text = '\n'.join(lines)
        else:
            return JsonResponse(
                {'error': f'Unsupported file type (.{ext}). Use PDF, DOCX, TXT, CSV, or XLSX.'},
                status=400,
            )
    except Exception as e:
        return JsonResponse({'error': f'Could not read file: {e}'}, status=400)

    if not text.strip():
        return JsonResponse({'error': 'No text could be extracted from the file.'}, status=400)

    if len(text) > 15000:
        text = text[:15000] + '\n[… file truncated for length …]'

    # MULTI-PERSON SUPPORT: append to any existing uploaded data rather than replace,
    # so the user can upload one document per employee and all data is preserved.
    separator = f'\n\n{"="*60}\nDOCUMENT: {filename}\n{"="*60}\n'
    if run.uploaded_data:
        combined = run.uploaded_data + separator + text
    else:
        combined = separator + text

    # Store up to 30 000 chars to handle large multi-person document sets
    run.uploaded_data = combined[:30000]
    run.save(update_fields=['uploaded_data'])

    # Tell the AI which file was just added so it can identify the person
    user_message = (
        f"I've uploaded a document for analysis: {filename}\n\n"
        f"Here is the content:\n\n{text}\n\n"
        "Please read this document carefully, identify the person it belongs to, "
        "extract their strengths, growth opportunities, and any relevant context, "
        "then summarise what you found and confirm with me before continuing."
    )

    try:
        ai_reply, suggestions = chat_with_ai(run, user_message)
    except Exception:
        ai_reply   = "I've received your file. Could you also tell me a bit about what you'd like me to focus on in this data?"
        suggestions = [
            "Look for skill gaps",
            "Identify training priorities",
            "Summarize the key themes",
            "Use this for my training plan",
        ]

    return JsonResponse({
        'filename':        filename,
        'display_message': f'📎 Uploaded: {filename}',
        'message':         ai_reply,
        'suggestions':     suggestions,
        'stage':           run.current_stage,
        'status':          run.status,
        'run_id':          run.id,
    })


@login_required
def load_thread(request, run_id):
    run     = get_object_or_404(BuilderRun, id=run_id, user=request.user)
    request.session['run_id'] = run.id
    history = run.get_chat_history()
    last_ai = next(
        (m['content'] for m in reversed(history) if m.get('role') in ('assistant', 'ai')),
        '',
    )
    suggestions = []
    if run.status not in ('plan_generated', 'cta_clicked'):
        suggestions = get_suggestions_for_message(last_ai)
    return JsonResponse({
        'type':        'resume',
        'history':     history,
        'stage':       run.current_stage,
        'status':      run.status,
        'run_id':      run.id,
        'title':       run.display_title(),
        'suggestions': suggestions,
    })


@login_required
@require_http_methods(["POST"])
def book(request, run_id):
    run = get_object_or_404(BuilderRun, id=run_id, user=request.user)
    try:
        data = json.loads(request.body)
        run.contact_name  = data.get('name',  '')
        run.contact_email = data.get('email', '')
        run.contact_phone = data.get('phone', '')
        run.booked_date   = data.get('date',  '')
        run.booked_time   = data.get('time',  '')
        run.status        = 'cta_clicked'
        run.save()

        # ── Send confirmation email ────────────────────────────────────────────
        from_email = getattr(settings, 'DEFAULT_FROM_EMAIL', 'noreply@bundle.com')

        # Email to the contact who booked
        if run.contact_email:
            subject_to_contact = 'Your Bundle Training Consultation is Confirmed!'
            body_to_contact = (
                f"Hi {run.contact_name or 'there'},\n\n"
                f"Thank you for booking a consultation with Bundle Training!\n\n"
                f"Here are your booking details:\n"
                f"  • Date: {run.booked_date or 'TBC'}\n"
                f"  • Time: {run.booked_time or 'TBC'}\n"
                f"  • Company: {run.company_name or 'N/A'}\n\n"
                f"We'll be in touch shortly to confirm the appointment and share "
                f"a preview of your personalised training plan.\n\n"
                f"If you need to make any changes, just reply to this email.\n\n"
                f"Warm regards,\n"
                f"The Bundle Training Team"
            )
            send_mail(
                subject_to_contact,
                body_to_contact,
                from_email,
                [run.contact_email],
                fail_silently=True,
            )

        # Internal notification email (if ADMIN_EMAIL is configured)
        admin_email = getattr(settings, 'ADMIN_NOTIFICATION_EMAIL', '')
        if admin_email:
            subject_internal = f'New Booking: {run.contact_name or "Unknown"} — {run.company_name or "Unknown company"}'
            body_internal = (
                f"A new consultation has been booked via the Bundle Training Builder.\n\n"
                f"Contact Details:\n"
                f"  • Name:  {run.contact_name or 'N/A'}\n"
                f"  • Email: {run.contact_email or 'N/A'}\n"
                f"  • Phone: {run.contact_phone or 'N/A'}\n\n"
                f"Booking Details:\n"
                f"  • Date:  {run.booked_date or 'TBC'}\n"
                f"  • Time:  {run.booked_time or 'TBC'}\n\n"
                f"Company Details:\n"
                f"  • Company:  {run.company_name or 'N/A'}\n"
                f"  • Industry: {run.industry or 'N/A'}\n"
                f"  • Team size: {run.team_size or 'N/A'}\n\n"
                f"View the full run in the admin panel: /admin/builder/builderrun/{run.id}/change/\n"
            )
            send_mail(
                subject_internal,
                body_internal,
                from_email,
                [admin_email],
                fail_silently=True,
            )

        return JsonResponse({'success': True, 'redirect': f'/confirm/{run.id}/'})
    except Exception:
        return JsonResponse({'error': 'Failed to save booking. Please try again.'}, status=500)


@login_required
def confirm(request, run_id):
    run = get_object_or_404(BuilderRun, id=run_id, user=request.user)
    return render(request, 'builder/confirmation.html', {'run': run})


@login_required
def download_pdf(request, run_id):
    run = get_object_or_404(BuilderRun, id=run_id, user=request.user)
    if not run.final_plan:
        return HttpResponse('No plan generated yet.', status=404)

    try:
        from .pdf_export import generate_plan_pdf
        pdf_bytes = generate_plan_pdf(run)
        company   = run.company_name or 'Training'
        filename  = f"Bundle_{company.replace(' ', '_')}_Training_Plan.pdf"
        response  = HttpResponse(pdf_bytes, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    except Exception as e:
        return HttpResponse(f'PDF generation error: {e}', status=500)
